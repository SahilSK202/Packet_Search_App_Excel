from enum import unique
from unicodedata import category
from flask import Flask, render_template, request, session
from datetime import datetime
import numpy as np
import pandas as pd
from os import path
import glob
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy

app = Flask(__name__)
app.secret_key = "123"
DEFAULT_FACTOR = 1500


@app.route('/', methods=['GET', 'POST'])
def Home():
    # Reading all packet files from data folder
    all_packets_path = ".\data\packets"

    filenames = glob.glob(all_packets_path + "\*.xlsx")

    available_packets = []
    available_packets_files = []
    for file in filenames:
        df = pd.read_excel(file, nrows=9, usecols=[
            'Key', 'Value'], )
        df.loc[~(df == 0).all(axis=1)]
        metadata = dict(df.fillna(0).values)
        # Filtering today available data
        if (type(metadata["End_Date"]) == str):
            metadata["End_Date"] = datetime.strptime(
                metadata["End_Date"], '%Y-%m-%d')

        if (metadata["End_Date"] > datetime.now()):
            available_packets.append(metadata["PacketNr"])
            available_packets_files.append(file)

    # Transforming for rendering purposes only
    today_available = '[' + ','. join(available_packets) + ']'

    if request.method == "POST":
        try:
            selected_pno = str(request.form['packet_list']).strip()
            # IF form is submitted
            if (selected_pno == 'Choose Packet'):

                packet_file_name = available_packets_files[available_packets.index(
                    session["OLD_PNO"])]
                new_end_date = str(request.form['userinput_endDate'])
                if (type(new_end_date) == str):
                    new_end_date = datetime.strptime(
                        new_end_date, '%Y-%m-%d')
                # Get the metadata
                metadata = pd.read_excel(packet_file_name, nrows=9, usecols=[
                                         'Key', 'Value'])
                metadata.loc[~(metadata == 0).all(axis=1)]
                metadata = dict(metadata.fillna(0).values)
                End_Date = metadata["End_Date"]
                if (type(metadata["End_Date"]) == str):
                    metadata["End_Date"] = datetime.strptime(
                        metadata["End_Date"], '%Y-%m-%d')

                # Check if date is updated , update the excel sheet
                if (str(End_Date)[0:10] != new_end_date):

                    # load excel file
                    workbook = load_workbook(packet_file_name)

                    # open workbook
                    sheet = workbook.active

                    # modify the desired cell

                    sheet.cell(row=6, column=2).value = new_end_date

                    # save the file
                    workbook.save(filename=packet_file_name)

                    session.pop('OLD_PNO', None)

            else:
                session["OLD_PNO"] = selected_pno

                filtered_data = pd.read_excel(available_packets_files[available_packets.index(
                    selected_pno)], header=10).fillna(0)
                metadata = pd.read_excel(available_packets_files[available_packets.index(
                    selected_pno)], nrows=9, usecols=['Key', 'Value'])
                # reading metadata of the packet
                metadata.loc[~(metadata == 0).all(axis=1)]
                metadata = dict(metadata.fillna(0).values)
                End_Date = metadata["End_Date"]
                if (type(metadata["End_Date"]) == str):
                    metadata["End_Date"] = datetime.strptime(
                        metadata["End_Date"], '%Y-%m-%d')
                all_columns = ['Chemicals']
                new_columns = []

                # Adding extra columns having name "_g"
                for col in filtered_data.columns[1:]:
                    filtered_data[col+'_g'] = ((DEFAULT_FACTOR /
                                                filtered_data[col].sum())*filtered_data[col]).round(2)
                    all_columns.append(col)
                    all_columns.append(col+'_g')
                    new_columns.append(col+'_g')

                filtered_data = filtered_data[all_columns]
                # Additon of columns
                addition = filtered_data[all_columns[1:]].sum(axis=0).T
                filtered_data.loc[len(filtered_data.index)] = [
                    'Sum', *addition.values]

                # # Setting default value for Factor
                default_factor_list = [1500/[*addition.values][i-1] if i % 2 != 0 else 0 for i in range(
                    len(filtered_data.columns)-1)]
                # print(len(filtered_data.columns)-1, default_factor_list)

                filtered_data.loc[len(filtered_data.index)] = [
                    'Factor'] + default_factor_list

                # filtered_data.loc[len(filtered_data.index)] = [
                #     'Factor'] + np.zeros(len(filtered_data.columns) - 1).tolist()

                # Collecting new column values to pass on frontend
                datavalues = np.transpose(
                    filtered_data[filtered_data.columns[1::2]].to_numpy()).tolist()

                column_names = '[' + ','. join(all_columns[1::2]) + ']'
                return render_template('index.html', today_available=today_available, End_Date=End_Date, datavalues=datavalues, column_names=column_names, selected_pno=selected_pno, data=filtered_data.round(2).to_html(index=False).replace('<td>', '<td style="text-align:center">').replace('<th>', '<th style="text-align:center">'))
        except Exception as e:
            print(e)
            return render_template('index.html', today_available=today_available)

    return render_template('index.html', today_available=today_available)


if __name__ == '__main__':

    app.run(debug=True, host="0.0.0.0", port=5000)
